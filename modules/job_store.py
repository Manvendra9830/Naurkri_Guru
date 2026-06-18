"""
Shared storage for scraped jobs from Indeed and future sources.
Writes to: all excels/manual_review_jobs.xlsx (and .csv)
"""

import os
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from modules.helpers import print_lg

MANUAL_REVIEW_CSV  = "all excels/manual_review_jobs.csv"
MANUAL_REVIEW_XLSX = "all excels/manual_review_jobs.xlsx"

SCHEMA = [
    "job_id", "title", "company", "location",
    "job_url", "source", "search_term",
    "scraped_at", "applied", "notes",
]


def _load_existing_ids():
    """Returns a set of (job_id, source) tuples already saved."""
    seen = set()
    if not os.path.exists(MANUAL_REVIEW_CSV):
        return seen
    try:
        with open(MANUAL_REVIEW_CSV, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                seen.add((row.get("job_id", ""), row.get("source", "")))
    except Exception as e:
        print_lg(f"[JOB-STORE] Could not read existing CSV: {e}")
    return seen


def save_scraped_job(record: dict) -> bool:
    """
    Saves a single scraped job record to the CSV and XLSX.
    Skips if the (job_id, source) pair already exists.
    Returns True if saved, False if skipped (duplicate).
    """
    existing = _load_existing_ids()
    key = (record.get("job_id", ""), record.get("source", ""))

    if key in existing:
        return False  # already saved, skip

    # Write to CSV
    file_exists = os.path.exists(MANUAL_REVIEW_CSV)
    try:
        with open(MANUAL_REVIEW_CSV, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=SCHEMA, extrasaction="ignore")
            if not file_exists:
                writer.writeheader()
            writer.writerow({k: record.get(k, "") for k in SCHEMA})
    except Exception as e:
        print_lg(f"[JOB-STORE] CSV write failed: {e}")
        return False

    # Rebuild XLSX from CSV (keeps it in sync)
    _rebuild_xlsx()
    return True


def _rebuild_xlsx():
    """Reads the full CSV and writes a clean, formatted XLSX."""
    try:
        rows = []
        if os.path.exists(MANUAL_REVIEW_CSV):
            with open(MANUAL_REVIEW_CSV, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                rows = list(reader)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Manual Review Jobs"

        # Header styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Column widths
        col_widths = {
            "job_id": 20, "title": 35, "company": 25, "location": 20,
            "job_url": 55, "source": 12, "search_term": 25,
            "scraped_at": 18, "applied": 10, "notes": 30,
        }

        # Write headers
        for col_idx, col_name in enumerate(SCHEMA, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = col_widths.get(col_name, 18)

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        # Source-based row colours
        source_colors = {
            "Indeed":      "EBF5FB",   # light blue
        }
        alt_fill_cache = {k: PatternFill("solid", fgColor=v) for k, v in source_colors.items()}
        default_fill   = PatternFill("solid", fgColor="FFFFFF")

        # Write data rows
        for row_idx, row in enumerate(rows, 2):
            source = row.get("source", "")
            row_fill = alt_fill_cache.get(source, default_fill)
            for col_idx, col_name in enumerate(SCHEMA, 1):
                value = row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                # Make job_url a clickable hyperlink
                if col_name == "job_url" and value:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")

        wb.save(MANUAL_REVIEW_XLSX)
        print_lg(f"[JOB-STORE] XLSX updated: {MANUAL_REVIEW_XLSX} ({len(rows)} jobs total)")

    except Exception as e:
        print_lg(f"[JOB-STORE] XLSX rebuild failed: {e}")


def get_all_scraped_jobs() -> list[dict]:
    """Returns all scraped jobs as a list of dicts."""
    if not os.path.exists(MANUAL_REVIEW_CSV):
        return []
    try:
        with open(MANUAL_REVIEW_CSV, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader)
    except Exception:
        return []
